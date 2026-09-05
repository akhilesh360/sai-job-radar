"""Aggregate DOL LCA disclosure files → per (employer, fiscal year) sponsorship stats. Streams each xlsx; low memory."""
import csv, glob, os, re, statistics, sys, time
import openpyxl
D = os.path.expanduser("~/h1b-data"); OUT = f"{D}/lca_agg.csv"
files = sorted(glob.glob(f"{D}/LCA_Disclosure_Data_FY20??_Q?.xlsx"), key=lambda f: (re.search(r"FY(\d{4})", f).group(1), re.search(r"_Q(\d)", f).group(1)))
DATA_ROLE = re.compile(r"data (?:engineer|scientist|analyst|architect|platform)|analytics|machine learning|\bml\b|\bai\b|artificial intelligence|business intelligence|\betl\b|big data|data warehouse|\bdatabricks\b|\bspark\b|statistic", re.I)
UNIT = {"Year": 1, "Month": 12, "Bi-Weekly": 26, "Week": 52, "Hour": 2080}
agg = {}      # (employer_raw, fy) → dict
seen = set(); seen_fy = None
t0 = time.time(); total = 0
for f in files:
    fy = int(re.search(r"FY(\d{4})", f).group(1))
    if fy != seen_fy: seen = set(); seen_fy = fy
    wb = openpyxl.load_workbook(f, read_only=True); ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True); header = [str(h).strip() if h else "" for h in next(rows)]
    ix = {h: i for i, h in enumerate(header)}
    g = lambda r, k: r[ix[k]] if k in ix and ix[k] < len(r) else None
    n = 0
    for r in rows:
        n += 1; total += 1
        if total % 250000 == 0: print(f"  {total:,} rows… ({time.time()-t0:.0f}s, {os.path.basename(f)})", file=sys.stderr, flush=True)
        if str(g(r, "VISA_CLASS") or "").strip() != "H-1B" or str(g(r, "CASE_STATUS") or "").strip() != "Certified": continue
        cn = str(g(r, "CASE_NUMBER") or "").strip()
        if not cn or cn in seen: continue
        seen.add(cn)
        emp = str(g(r, "EMPLOYER_NAME") or "").strip()
        if not emp: continue
        key = (emp, fy); a = agg.get(key)
        if a is None: a = agg[key] = {"lcas": 0, "positions": 0, "data_lcas": 0, "data_wages": [], "states": {}, "titles": {}}
        a["lcas"] += 1
        try: a["positions"] += int(g(r, "TOTAL_WORKER_POSITIONS") or 1)
        except (TypeError, ValueError): a["positions"] += 1
        st = g(r, "WORKSITE_STATE") or g(r, "EMPLOYER_STATE")
        if st: a["states"][st] = a["states"].get(st, 0) + 1
        title = f"{str(g(r, 'JOB_TITLE') or '')} {str(g(r, 'SOC_TITLE') or '')}"
        if DATA_ROLE.search(title):
            a["data_lcas"] += 1
            t = str(g(r, "JOB_TITLE") or "").strip()[:60]
            if t: a["titles"][t] = a["titles"].get(t, 0) + 1
            try:
                w = float(g(r, "WAGE_RATE_OF_PAY_FROM") or 0) * UNIT.get(str(g(r, "WAGE_UNIT_OF_PAY") or "Year").strip(), 1)
                if 20000 <= w <= 1_000_000: a["data_wages"].append(w)
            except (TypeError, ValueError): pass
    wb.close()
    print(f"{os.path.basename(f)}: {n:,} rows | certified H-1B cases so far FY{fy}: {len(seen):,} | employers: {len(agg):,} | {time.time()-t0:.0f}s", file=sys.stderr, flush=True)
    # write after every file so a crash still leaves usable output
    with open(OUT, "w", newline="") as fh:
        w = csv.writer(fh); w.writerow(["employer", "fy", "lcas", "positions", "data_lcas", "data_wage_p25", "data_wage_median", "data_wage_p75", "top_states", "top_data_titles"])
        for (emp, y), a in agg.items():
            dw = sorted(a["data_wages"]); q = (lambda p: int(dw[min(len(dw)-1, int(len(dw)*p))])) if dw else (lambda p: "")
            states = ",".join(k for k, _ in sorted(a["states"].items(), key=lambda kv: -kv[1])[:3])
            titles = " | ".join(k for k, _ in sorted(a["titles"].items(), key=lambda kv: -kv[1])[:3])
            w.writerow([emp, y, a["lcas"], a["positions"], a["data_lcas"], q(0.25), q(0.5), q(0.75), states, titles])
print(f"DONE: {total:,} rows read, {len(agg):,} employer-years → {OUT} in {time.time()-t0:.0f}s", file=sys.stderr)
